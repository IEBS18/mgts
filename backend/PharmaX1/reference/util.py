import os
import re
import requests
from openpyxl import load_workbook, Workbook
from typing import Union, List


# class ExcelPreProcessor:
#     def __init__(self, filepath):
#         self.filepath = filepath
#         self.wb = load_workbook(filepath)
#         self.ws = self.wb.active
    
#     def add_unique_key(self):
#         file_name = os.path.splitext(os.path.basename(self.filepath))[0]
#         print(file_name)
#         headers = [cell.value for cell in self.ws[1]]
#         print(headers)
#         if 'unique_key' in headers:
#             print("The 'unique_key' column already exists in the file.")
#         else:
#             unique_key_col = len(headers) + 1
#             self.ws.cell(row=1, column=unique_key_col, value='unique_key')
#             for i in range(2, self.ws.max_row + 1):
#                 self.ws.cell(row=i, column=unique_key_col, value=f"{file_name}_{i-1}")
#             print("The 'unique_key' column has been added successfully.")

#     def save(self):
#         self.wb.save(self.filepath)


class ExcelFileProcessor:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = load_workbook(filename=file_path)
        self.sheet = self.workbook.active

        self.defaultKeys = ['unique_key','Organization_Name', 'Product_Name', 'Ingredients', 'Routes_of_Administration', 'Warnings','Clinical_Pharmacology','Indications_and_Usage', 'Category' ]

    def find_column_by_header(self, header):
        """Finds and returns the column number of a given header."""
        for cell in self.sheet[1]:
            if cell.value == header:
                return cell.column
        return None

    def find_first_empty_column(self):
        """Finds and returns the first empty column."""
        for cell in self.sheet[1]:
            if cell.value is None:
                return cell.column
        return self.sheet.max_column + 1

    def find_first_empty_row_in_column(self, column):
        """Finds and returns the first empty row number in a specific column."""
        for row in range(2, self.sheet.max_row + 2):
            if self.sheet.cell(row=row, column=column).value is None:
                return row
        return self.sheet.max_row + 1
    
    def ensure_unique_key_column(self):
        """Ensures that the 'unique_key' column exists, creating it if necessary."""
        unique_key_col = self.find_column_by_header('unique_key')
        if unique_key_col is None:
            empty_col = self.find_first_empty_column()
            self.sheet.cell(row=1, column=empty_col, value='unique_key')
            for i in range(2, self.sheet.max_row + 1):
                self.sheet.cell(row=i, column=empty_col, value=f"{os.path.splitext(os.path.basename(self.file_path))[0]}_{i-1}")
            self.workbook.save(self.file_path)
            print("The 'unique_key' column has been added successfully.")
        else:
            print("The 'unique_key' column already exists in the file.")
        
    
    def extractor(self, keys : Union[List[str], str] = None):
        """
        Extracts specific data from the Excel file and returns it as a pandas DataFrame.

        Parameters:
        -----------
        keys : Union[List[str], str], optional
            A list of column names or a single column name to extract from the Excel file. 
            If 'all', all columns will be extracted. If None, the default keys specified 
            in `self.defaultKeys` will be used. Default is None.

        Returns:
        --------
        pd.DataFrame
            A pandas DataFrame containing the extracted data with columns in lowercase.

        Raises:
        -------
        KeyError
            If any of the specified keys are not found in the header of the Excel sheet.

        Example:
        --------
        >>> processor = ExcelFileProcessor('path/to/excel/file.xlsx')
        >>> df_all = processor.extractor(keys="all")
        >>> print(df_all)
        name  age      city
        0  John   30  New York
        1  Jane   25    Boston
        2   Doe   22   Chicago
        
        >>> df_selected = processor.extractor(keys=["Name", "City"])
        >>> print(df_selected)
        name      city
        0  John  New York
        1  Jane    Boston
        2   Doe   Chicago
        """
        data = []
        header = [cell.value for cell in self.sheet[1]]
        col_index = {name: index for index, name in enumerate(header, start=1)}

        if keys == "all":
            keys = header
        elif not keys:
            keys = self.defaultKeys

        available_cols = {key: col_index.get(key) for key in keys if key in col_index}
        print("using headers :", available_cols)
        
        data = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if any(row):
                row_dict = {}
                for key, index in available_cols.items():
                    row_dict[key.lower()] = row[index - 1] if index is not None else None
                data.append(row_dict)
        
        # Convert to DataFrame
        # df = pd.DataFrame(data, columns=available_cols.keys())
        
        return data

    def row_extractor(self, row_num, keys : Union[List[str], str] = None ):
        """Extracts data from a specific row in the Excel file."""
        if row_num < 2 or row_num > self.sheet.max_row:
            raise ValueError("Invalid row number. It should be between 2 and the maximum row number.")
        
        header = [cell.value for cell in self.sheet[1]]
        col_index = {name: index for index, name in enumerate(header, start=1)}

        if keys == "all":
            keys = header
        elif not keys:
            keys = self.defaultKeys
        available_cols = {key: col_index.get(key) for key in keys if key in col_index}

        row_data = []
        for row in self.sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True):
            if any(row):
                row_dict = {}
                for key, index in available_cols.items():
                    row_dict[key.lower()] = row[index - 1] if index is not None else None
                row_data.append(row_dict)
        
        return row_data

if __name__ == '__main__':
    processor = ExcelFileProcessor('data.xlsx')
    processor.ensure_unique_key_column()
    